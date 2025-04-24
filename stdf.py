import copy
import os
from enum import Enum, auto
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
import numpy as np
import matplotlib.pyplot as plt
from tkinter import messagebox
import warnings
from fpdf import FPDF
from matplotlib.ticker import FormatStrFormatter
from pikepdf import Pdf, OutlineItem
from wafer_map import generate_wafer_map
from PIL import Image

# warnings.filterwarnings('ignore', category=RuntimeWarning)
plt.switch_backend('agg')
FILE_INFO = "File Info"
DUT_SUMMARY = "DUT Summary"


class FilterType(Enum):
    ALL_RUN = 1
    PASSED_RUN = 2
    ONE_SIGMA = 3
    TWO_SIGMA = 4
    THREE_SIGMA = 5
    SIX_SIGMA = 6
    NONE = 7

def from_enum_to_string(enum):
    if enum == FilterType.ALL_RUN or enum == FilterType.NONE:
        return "All Runs"
    elif enum == FilterType.PASSED_RUN:
        return "Pass Runs"
    elif enum == FilterType.ONE_SIGMA:
        return "1\u03C3"
    elif enum == FilterType.TWO_SIGMA:
        return "2\u03C3"
    elif enum == FilterType.THREE_SIGMA:
        return "3\u03C3"
    else:
        return "6\u03C3"


class StatisticalResults:
    def __init__(self):
        self.min = 0
        self.max = 0
        self.mean = 0
        self.median = 0
        self.range = 0
        self.std = 0
        self.cp = 0
        self.cpk = 0


class coordinate:
    def __init__(self, x=0, y=0):
        self.x = x
        self.y = y


class TestInfo:
    def __init__(self):
        self.test_description = ""
        self.lower_limit = 0
        self.upper_limit = 0
        self.unit = ""
        self.x = np.array([], dtype=np.int8)
        self.y = np.array([], dtype=np.int8)
        self.is_passed = np.array([], dtype=object)
        self.values = np.array([], dtype=np.float64)
        self.stats = StatisticalResults()
        self.total_run = 0
        self.dropped = 0

    def generate_statistics(self):

        self.stats.min = np.nanmin(self.values)
        self.stats.max = np.nanmax(self.values)
        self.stats.mean = np.nanmean(self.values)
        self.stats.std = 0.0 if np.nanstd(self.values) <= 1.0e-6 else np.nanstd(self.values)
        self.stats.range = self.stats.max - self.stats.min
        self.stats.median = np.nanmedian(self.values)

        try:
            self.stats.cp = (self.upper_limit - self.lower_limit) / (6. * self.stats.std)
            self.stats.cpk = min((self.upper_limit - self.stats.mean) / (3 * self.stats.std),
                                 (self.stats.mean - self.lower_limit) / (3 * self.stats.std))
        except:
            self.stats.cp = 9999.999
            self.stats.cpk = 9999.999

        # with warnings.catch_warnings():
        #     warnings.filterwarnings('error')
        #     try:
        #         self.stats.cp = (self.upper_limit - self.lower_limit) / (6.*self.stats.std)
        #         self.stats.cpk = min((self.upper_limit - self.stats.mean) / (3 * self.stats.std),
        #                              (self.stats.mean - self.lower_limit) / (3 * self.stats.std))
        #     except Warning as e:
        #
        #         # print(self.test_description, self.unit, f'std: { self.stats.std}')
        #         # multiplier = 1 if self.unit != "HEX" else -1
        #         self.stats.cp = 9999.999
        #         self.stats.cpk = 9999.999
        # if self.stats.std <= 1.0e-5:
        #     self.stats.std = 0.0


class WaferRunInfo:
    def __init__(self):
        self.family_id = ""
        self.product_id = ""
        self.start_time = ""
        self.os = ""
        self.tester = ""
        self.lot = ""
        self.sites = ""
        self.passed_runs = ""
        self.all_runs = ""




def query_wafer_run_info(file_name):
    info_book = load_workbook(file_name)
    File_Info_Sheet = info_book["File Info"]
    max_row = File_Info_Sheet.max_row
    info_dict = {}

    for row_number in range(1, max_row+1):
        # print(File_Info_Sheet["A"+str(row_number)].value)
        info_dict[File_Info_Sheet["A"+str(row_number)].value.strip()] = File_Info_Sheet["B"+str(row_number)].value

    wafer_run_info_ = WaferRunInfo()
    wafer_run_info_.passed_runs = info_dict["DUTs Passed:"]
    wafer_run_info_.all_runs = str(int(info_dict["DUTs Passed:"]) + int(info_dict["DUTs Failed:"]))
    wafer_run_info_.family_id = info_dict["Family ID:"]
    wafer_run_info_.product_id = info_dict["Product ID:"]
    wafer_run_info_.start_time = info_dict["Start Time:"]
    wafer_run_info_.os = info_dict["Tester Software Version:"]+" / "+info_dict["Tester Software Type:"]
    wafer_run_info_.tester = info_dict["Tester Type:"]+" / "+info_dict["Node Name:"]
    wafer_run_info_.lot = info_dict["Lot ID:"]
    wafer_run_info_.sites = info_dict["Station Number:"]
    return wafer_run_info_


def query_wafer_run_info_(sheet):
    max_row = sheet.max_row
    info_dict = {}

    for row_number in range(1, max_row+1):
        info_dict[sheet["A"+str(row_number)].value.strip()] = sheet["B"+str(row_number)].value

    wafer_run_info_ = WaferRunInfo()
    wafer_run_info_.passed_runs = info_dict["DUTs Passed:"]
    wafer_run_info_.all_runs = str(int(info_dict["DUTs Passed:"]) + int(info_dict["DUTs Failed:"]))
    # print(wafer_run_info_.all_runs)
    wafer_run_info_.family_id = info_dict["Family ID:"]
    wafer_run_info_.product_id = info_dict["Product ID:"]
    wafer_run_info_.start_time = info_dict["Start Time:"]
    wafer_run_info_.os = info_dict["Tester Software Version:"]+" / "+info_dict["Tester Software Type:"]
    wafer_run_info_.tester = info_dict["Tester Type:"]+" / "+info_dict["Node Name:"]
    wafer_run_info_.lot = info_dict["Lot ID:"]
    wafer_run_info_.sites = info_dict["Station Number:"]
    return wafer_run_info_


def generate_stats_report(import_file, filter_type, export_file, progress_bar, generate_button, progress_label=None, src_dict={}):
    generate_button.state(["disabled"])
    title = from_enum_to_string(filter_type)
    wook_book = load_workbook(import_file)
    DUT_Data_Sheet = wook_book[DUT_SUMMARY]
    File_Info_Sheet = wook_book[FILE_INFO]
    wafer_run_info = query_wafer_run_info_(File_Info_Sheet)

    max_row = DUT_Data_Sheet.max_row
    max_column = DUT_Data_Sheet.max_column
    column_list = []
    for index in range(12, max_column + 1):
        column_list.append(get_column_letter(index))
    test_dict = {}
    total_wafer_run = False

    for col in column_list:
        if DUT_Data_Sheet[col + str(3)].value != "N/A":
            test_id = int(DUT_Data_Sheet[col + str(2)].value)  # Test Number
            test_info = TestInfo()
            test_info.test_description = DUT_Data_Sheet[col + str(1)].value.strip()
            test_info.lower_limit = float(DUT_Data_Sheet[col + str(4)].value)  # LLimit
            test_info.upper_limit = float(DUT_Data_Sheet[col + str(3)].value)  # HLimit
            test_info.unit = DUT_Data_Sheet[col + str(5)].value  # Unit
            values = []
            x = []
            y = []
            temp_passed = []
            total_wafer_run_passed = []
            passed_run = []
            for row_number in range(6, max_row + 1):
                coord = DUT_Data_Sheet["J" + str(row_number)].value
                x.append(int(coord.split()[0][1:-1]))
                y.append(int(coord.split()[1][0:-1]))
                is_failed = DUT_Data_Sheet["K" + str(row_number)].value
                if is_failed == "Failed - 0x08":
                    total_wafer_run_passed.append(False)
                    temp_passed.append(False)
                    passed_run.append(False)
                    # if filter_type == FilterType.PASSED_RUN:
                    #     continue
                else:
                    total_wafer_run_passed.append(True)
                    temp_passed.append(True)
                    passed_run.append(True)
                cell_value = DUT_Data_Sheet[col + str(row_number)].value
                try:
                    values.append(float(cell_value))
                    if test_info.lower_limit <= cell_value <= test_info.upper_limit:
                        temp_passed[-1] = True
                except:
                    values.append(np.nan)
                    temp_passed[-1] = None

            test_info.x = np.array(x)
            test_info.y = np.array(y)
            test_info.is_passed = np.array(temp_passed, dtype=object)
            total_passed_array = np.array(total_wafer_run_passed)
            passed_array = np.array(passed_run, dtype=bool)


            # reverse y and shift x and y to coordinate 0, 0
            x_min = min(test_info.x)
            y_shift = max(test_info.y)
            test_info.x = test_info.x + abs(x_min)
            # test_info.y = y_shift - test_info.y
            # Thread(
            #     target=generate_wafer_map,
            #     args=(test_info.x, test_info.y, y_shift, test_info.is_passed, str(test_id)+"_wafermap.png",
            #           from_enum_to_string(filter_type)+"_wafer_map/"),
            #     daemon=True
            # ).start()

            if not total_wafer_run:
                generate_wafer_map(test_info.x, test_info.y, y_shift, total_passed_array,
                                   f'{wafer_run_info.family_id} Wafer Run',
                                   f'{wafer_run_info.family_id}_Wafer_Run.png',
                                   f'{wafer_run_info.family_id} Wafer Map/')

            total_wafer_run = True

            title_ = f'Test ID: {test_id} Wafer Map' if (filter_type == FilterType.ALL_RUN or
                                                         filter_type == FilterType.PASSED_RUN) \
                else f'Test ID: {test_id} before {title} filter'
            generate_wafer_map(test_info.x, test_info.y, y_shift, test_info.is_passed, title_,
                               str(test_id)+"_before_wafermap.png",
                               title + " Wafer Map/")

            test_info.values = np.array(values)
            value_array_backup = copy.copy(test_info.values)
            # print(test_info.values, len(test_info.values))
            if filter_type == FilterType.PASSED_RUN:
                # passed_index = np.where(passed_array == 1)
                # print(passed_index, len(passed_index))
                test_info.values = test_info.values[passed_array]
            # print(test_info.values)

            no_nan_mask = ~np.isnan(value_array_backup)
            # nan_mask = np.isnan(value_array_backup)
            test_info.total_run = len(value_array_backup[no_nan_mask])
            if filter_type == FilterType.PASSED_RUN:
                test_info.dropped = (int(wafer_run_info.all_runs) - int(wafer_run_info.passed_runs) -
                                     len(value_array_backup[~no_nan_mask]))
                # test_info.total_run = int(wafer_run_info.all_runs)
                # test_info.dropped = int(wafer_run_info.all_runs) - int(wafer_run_info.passed_runs)
                # print(test_info.dropped, test_info.total_run)
            elif filter_type == FilterType.ALL_RUN:
                # test_info.total_run = int(wafer_run_info.all_runs)
                test_info.dropped = 0
            test_info.generate_statistics()

            # select data based on filter
            # for filter type three sigma
            # if test_info.unit != "HEX" and test_info.stats.std > 0.001:

            if test_info.stats.std > 0.0001:
                mask = np.array([], dtype=bool)
                if filter_type != FilterType.ALL_RUN and filter_type != FilterType.PASSED_RUN:
                    # filter one sigma
                    if filter_type == FilterType.ONE_SIGMA:
                        mask = (((test_info.stats.mean - test_info.stats.std) <= test_info.values) &
                                (test_info.values <= (test_info.stats.mean + test_info.stats.std)))

                    # filter two sigma
                    elif filter_type == FilterType.TWO_SIGMA:
                        mask = (((test_info.stats.mean - 2 * test_info.stats.std) <= test_info.values) &
                                (test_info.values <= (test_info.stats.mean + 2 * test_info.stats.std)))

                    # filter three sigma
                    elif filter_type == FilterType.THREE_SIGMA:
                        mask = (((test_info.stats.mean - 3 * test_info.stats.std) <= test_info.values) &
                                (test_info.values <= (test_info.stats.mean + 3 * test_info.stats.std)))

                    # filter six sigma
                    elif filter_type == FilterType.SIX_SIGMA:
                        mask = (((test_info.stats.mean - 6 * test_info.stats.std) <= test_info.values) &
                                (test_info.values <= (test_info.stats.mean + 6 * test_info.stats.std)))

                    # update pass array and value array
                    index_array, = np.where(~mask)
                    for i in index_array:
                        if test_info.is_passed[i]:
                            test_info.is_passed[i] = np.nan
                    test_info.values = test_info.values[mask]
                    test_info.dropped = test_info.total_run - len(test_info.values)

                    # recalculate stats result after filter
                    if len(test_info.values) > 1:
                        test_info.generate_statistics()

                    # generate wafer map after filter
                    generate_wafer_map(test_info.x, test_info.y, y_shift, test_info.is_passed, f'Test ID: {test_id} '
                                                                                               f' after {title} filter',
                                       str(test_id) + "_after_wafermap.png",
                                       title + " Wafer Map/")

                    # combine two wafer map into one
                    images = [Image.open(title + "Wafer Map/"+str(test_id) + "_before_wafermap.png"),
                              Image.open(title + "Wafer Map/"+str(test_id) + "_after_wafermap.png")]
                    widths, heights = zip(*(i.size for i in images))

                    total_width = sum(widths)
                    max_height = max(heights)

                    new_im = Image.new('RGB', (total_width, max_height))
                    x_offset = 0
                    for im in images:
                        new_im.paste(im, (x_offset, 0))
                        x_offset += im.size[0]
                    # save to local dir
                    new_im.save(title + " Wafer Map/"+str(test_id) + "_before_after_wafermap.png")

            if test_id not in test_dict:
                test_dict[test_id] = test_info

    cpk_9999 = []

    for key, value in test_dict.items():
        # print(f'{value.test_description}: cp: {value.stats.cp}, cpk: {value.stats.cpk}')
        if value.stats.cp == -9999.999:
            cpk_9999.append([key, test_dict[key]])

    sorted_ = {k: v for k, v in sorted(test_dict.items(), key=lambda item: item[1].stats.cpk)}
    cpk_first_n = list(sorted_.items())[:40]

    statistics_report = "Statistics Report"

    pdf = FPDF()

    # add title
    pdf.add_page()
    pdf.set_font("helvetica", size=40, style="IB")
    pdf.set_text_color(255, 0, 0)
    pdf.cell(195, 70, txt=statistics_report, ln=1, align="C")

    pdf.set_font("helvetica", size=20, style="IB")
    pdf.set_text_color(0, 0, 250)
    pdf.cell(195, 10, txt=wafer_run_info.family_id, ln=1, align="C")
    #
    pdf.set_font("helvetica", size=16, style="B")
    pdf.set_text_color(0, 0, 0)
    pdf.cell(10, 10, txt=str(" " * 40 + str(wafer_run_info.product_id)), ln=1, align="L")
    #
    pdf.set_font("helvetica", size=16, style="B")
    pdf.set_text_color(0, 0, 0)
    pdf.cell(10, 10, txt=" " * 40 + wafer_run_info.start_time, ln=1, align="L")
    #
    pdf.set_font("helvetica", size=16, style="B")
    pdf.set_text_color(0, 0, 0)
    pdf.cell(10, 10, txt=" " * 108 + "OS: " + wafer_run_info.os, ln=1, align="C")
    #
    pdf.set_font("helvetica", size=16, style="B")
    pdf.set_text_color(0, 0, 0)
    pdf.cell(10, 10, txt=" " * 137 + "Tester: " + wafer_run_info.tester, ln=1, align="C")
    #
    pdf.set_font("helvetica", size=16, style="B")
    pdf.set_text_color(0, 0, 0)
    pdf.cell(10, 10, txt=" " * 96 + "Lot: " + str(wafer_run_info.lot), ln=1, align="C")
    #
    pdf.set_font("helvetica", size=16, style="B")
    pdf.set_text_color(0, 0, 0)
    pdf.cell(10, 10, txt=" " * 88 + "Sites: " + str(wafer_run_info.sites), ln=1, align="C")
    #
    pdf.set_font("helvetica", size=16, style="B")
    pdf.set_text_color(0, 0, 0)

    num_runs = wafer_run_info.passed_runs if filter_type == FilterType.PASSED_RUN else wafer_run_info.all_runs
    run_info = "All Runs: " if filter_type != FilterType.PASSED_RUN else "Passed Runs: "
    spacing = 100 if filter_type != FilterType.PASSED_RUN else 108
    pdf.cell(10, 10, txt=" " * spacing + run_info + str(num_runs), ln=1, align="C")

    pdf.image(f'{wafer_run_info.family_id} Wafer Map/{wafer_run_info.family_id}_Wafer_Run.png',
              x=45, y=170, w=120, h=110)

    table_data = [["TNUM", "SITE", "RANGE", "STD", "CP", "CPK", "UNITS", "TNAME"]]
    link_dic = {}
    for k, v in cpk_first_n:
        temp = []
        temp.append(str(k))
        temp.append(1)
        temp.append(round(v.stats.range, 3))
        temp.append(round(v.stats.std, 3))
        temp.append(round(v.stats.cp, 3))
        temp.append(round(v.stats.cpk, 3))
        temp.append(v.unit)
        temp.append(v.test_description)
        link_dic[v.test_description] = pdf.add_link()
        table_data.append(temp)

    pdf.add_page()
    pdf.set_font("helvetica", size=20, style="B")
    pdf.set_text_color(0, 0, 0)
    pdf.cell(10, 10, txt=" " * 85 + "Cpk Summary", ln=1, align="C")
    #
    pdf.ln(5)
    pdf.set_font('Helvetica', 'b', 8)
    pdf.set_text_color(0, 0, 0)
    row = 0
    col = 0

    for row_index, data_row in enumerate(table_data):
        for column_index, data in enumerate(data_row):
            if column_index != 7:
                if row_index == 0 or column_index != 0:
                    pdf.cell(w=18, h=6, txt=str(data), border=1, ln=0, align="R")
                else:
                    pdf.cell(w=18, h=6, txt=str(data), border=1, ln=0, align="R", link=link_dic[data_row[7]])
            else:
                if row_index == 0:
                    pdf.cell(w=54, h=6, txt=str(data), border=1, ln=0, align="R")
                else:
                    pdf.cell(w=54, h=6, txt=str(data), border=1, ln=0, align="R", link=link_dic[data])
        pdf.ln()

    # add trend plot and histogram
    plt.style.use('ggplot')
    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(10, 8), constrained_layout=True)
    # fig.subplots_adjust(top=0.9)
    # fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(8, 6))

    trend_title = "Trend Plot"
    histogram = "Histogram"
    index = 0
    load = len(test_dict)

    for test_id, test_info in test_dict.items():
        # print(test_id)
        test_description = test_info.test_description
        min_ = round(test_info.stats.min, 3)
        max_ = round(test_info.stats.max, 3)
        mean = round(test_info.stats.mean, 3)
        median = round(test_info.stats.median, 3)
        range_ = round(test_info.stats.range, 3)
        std = round(test_info.stats.std, 3)
        cp = round(test_info.stats.cp, 3)
        cpk = round(test_info.stats.cpk, 3)
        unit = test_info.unit
        l_lim = test_info.lower_limit
        h_lim = test_info.upper_limit
        number_points = len(test_info.values)
        number_array = np.arange(1, number_points + 1)
        value_array = test_info.values
        ax1.scatter(number_array, value_array, c="blue", s=2, marker="o")
        ax1.plot(number_array, value_array, "r-", linewidth=1)
        ax1.plot(number_array, np.repeat(h_lim, number_points), linewidth=2, linestyle="--", label=f"HLIM: {h_lim}")
        ax1.plot(number_array, np.repeat(l_lim, number_points), linewidth=2, linestyle="--", label=f"LLIM: {l_lim}")
        ax1.legend(loc="upper right")
        ax1.set_xlabel("runs")
        limit_gap = ((h_lim - l_lim) / 5.) * 0.2
        ax1.set_ylabel(unit)
        ax1.set_ylim(l_lim - limit_gap, h_lim + limit_gap)
        ax1.set_title(trend_title)

        ax2.axvline(x=l_lim, color="blue", linestyle="-", label=f"ll={l_lim}")
        ax2.axvline(x=h_lim, color="blue", linestyle="-", label=f"ul={h_lim}")
        ax2.axvline(x=median, color="grey", linestyle="--", label=f"median={median}")

        if filter_type == FilterType.ONE_SIGMA or filter_type == FilterType.PASSED_RUN or filter_type == FilterType.ALL_RUN:
            ax2.axvline(x=mean - std, color="green", linestyle="-",
                        label=f"{from_enum_to_string(FilterType.ONE_SIGMA)} ll={round(mean - std, 3)}")
            ax2.axvline(x=mean + std, color="green", linestyle="-",
                        label=f"{from_enum_to_string(FilterType.ONE_SIGMA)} ul={round(mean + std, 3)}")
        elif filter_type == FilterType.TWO_SIGMA:
            ax2.axvline(x=mean - 2 * std, color="green", linestyle="-",
                        label=f"{title} ll={round(mean- 2 * std, 3)}")
            ax2.axvline(x=mean + 2 * std, color="green", linestyle="-",
                        label=f"{title} ul={round(mean + 2 * std, 3)}")
        elif filter_type == FilterType.THREE_SIGMA:
            ax2.axvline(x=mean - 3 * std, color="green", linestyle="-",
                        label=f"{title} ll={round(mean - 3 * std, 3)}")
            ax2.axvline(x=mean + 3 * std, color="green", linestyle="-",
                        label=f"{title} ul={round(mean + 3 * std, 3)}")
        elif filter_type == FilterType.SIX_SIGMA:
            ax2.axvline(x=mean - 6 * std, color="green", linestyle="-",
                        label=f"{title} ll={round(mean - 6 * std, 3)}")
            ax2.axvline(x=mean + 6 * std, color="green", linestyle="-",
                        label=f"{title} ul={round(mean + 6 * std, 3)}")

        ax2.hist(value_array, bins=10, color="skyblue", edgecolor="black", align='mid')

        ax2.set_xlabel(unit)
        ax2.tick_params(axis='x', labelrotation=45)
        ax2.xaxis.set_major_formatter(FormatStrFormatter('%.2f'))
        ax2.set_title(histogram)
        ax2.legend(loc="upper right")
        plt.tight_layout()
        # fig.suptitle(str(test_id) + ": " + test_info.test_description, fontsize="x-large")
        fig.suptitle(f"Test {test_id} - {test_info.test_description}\n"
                     f"fallout = {test_info.dropped} out of {test_info.total_run} "
                     f"({round((test_info.dropped/test_info.total_run)*100, 2)}%)\n\u03BC = {mean} \u03C3 = {std} cpk = {round(cpk, 1)}\n"
                     f"dropped = {test_info.dropped}",
                     fontsize=14)
        temp = ""
        if filter_type == FilterType.PASSED_RUN:
            temp = "passed_runs_"
        elif filter_type == FilterType.ALL_RUN:
            temp = "all_runs_"
        elif filter_type == FilterType.ONE_SIGMA:
            temp = "1_sigma_runs_"
        elif filter_type == FilterType.TWO_SIGMA:
            temp = "2_sigma_runs_"
        elif filter_type == FilterType.THREE_SIGMA:
            temp = "3_sigma_runs_"
        else:
            temp = "6_sigma_runs_"

        directory = temp + "Plots/"
        if not os.path.exists(directory):
            os.makedirs(directory, exist_ok=True)

        wafer_map_file_name = f"{title} Wafer Map/{test_id}_before_after_wafermap.png"
        plot_file_name = directory + str(test_id) + ".png"
        fig.savefig(plot_file_name)
        ax1.cla()
        ax2.cla()

        pdf.add_page()
        if test_description in link_dic:
            pdf.set_link(link_dic[test_description])
        pdf.add_font(fname='DejaVuSansCondensed.ttf')
        pdf.set_font("helvetica", style="B", size=15)
        pdf.set_text_color(0, 0, 0)
        pdf.cell(180, 10, txt=f"Test: {test_id}", ln=1, align="C")
        pdf.set_font("helvetica", style="B", size=20)
        pdf.set_text_color(0, 0, 0)
        pdf.cell(180, 10, txt=f"{test_description}", ln=1, align="C")
        pdf.ln(3)
        pdf.set_font("DejaVuSansCondensed", size=10)
        pdf.set_text_color(0, 0, 0)
        parameter_table = [["SITE", "FILTER", "MIN", "MEAN", "MAX", "RANGE", "STD", "CP", "CPK", "UNITS"],
                           [wafer_run_info.sites, title, min_, mean, max_, range_, std, cp, cpk, unit]]

        for data_row in parameter_table:
            for row_index, data in enumerate(data_row):
                pdf.cell(w=19, h=6, txt=" " * 4 + str(data), border=1, ln=0, align="R")
            pdf.ln()

        # insert wafer map
        if filter_type != FilterType.ALL_RUN and filter_type != FilterType.PASSED_RUN:
            pdf.image(f"{wafer_map_file_name}", x=6, y=55, w=200, h=95)
        else:
            pdf.image(f"{title} Wafer Map/{test_id}_before_wafermap.png", x=45, y=55, w=120, h=110)

        # insert plot
        pdf.image(f"{plot_file_name}", x=0, y=170, w=200, h=100)
        index += 1
        progress_val = index/float(load)*100
        progress_bar["value"] = progress_val
        if progress_label is not None:
            progress_label.config(text=f"{progress_val:.0f}%")
        progress_bar.update_idletasks()

    pdf.output(export_file)

    if "src_path" not in src_dict:
        src_dict["src_path"] = []
    src_dict["src_path"].append(export_file)

    # add outline of pdf
    pdf_ = Pdf.open(export_file)
    outline_list = []
    with pdf_.open_outline() as outline:
        oi = OutlineItem("Cover page", 0)
        outline_list.append(oi)
        oi = OutlineItem("Cpk summary", 1)
        outline_list.append(oi)
        page_index = 1
        for test_id, test_info in test_dict.items():
            page_index += 1
            oi = OutlineItem(str(test_id) + ": " + test_info.test_description, page_index)
            outline_list.append(oi)
        outline.root.extend(outline_list)

    # add link page
    if not os.path.exists("../wafer run stats report tool/statistics report/"):
        os.makedirs("../wafer run stats report tool/statistics report/", exist_ok=True)
    pdf_.save(os.getcwd() + "/statistics report/" + os.path.basename(export_file))

    messagebox.showinfo("Completion!", "Generate stats report successfully!")
    generate_button.state(["!disabled"])


def generate_stats_reports_with_filter(import_file, move_keep_bar1_left, move_keep_bar2_right, filter_type, export_file, progress_bar, progress_label=None, src_dict={}):
    title = from_enum_to_string(filter_type)
    wook_book = load_workbook(import_file)
    DUT_Data_Sheet = wook_book[DUT_SUMMARY]
    File_Info_Sheet = wook_book[FILE_INFO]
    wafer_run_info = query_wafer_run_info_(File_Info_Sheet)

    max_row = DUT_Data_Sheet.max_row
    max_column = DUT_Data_Sheet.max_column
    column_list = []
    for index in range(12, max_column + 1):
        column_list.append(get_column_letter(index))
    test_dict = {}
    total_wafer_run = False

    for col in column_list:
        if DUT_Data_Sheet[col + str(3)].value != "N/A":
            test_id = int(DUT_Data_Sheet[col + str(2)].value)  # Test Number
            test_info = TestInfo()
            test_info.test_description = DUT_Data_Sheet[col + str(1)].value.strip()
            test_info.lower_limit = float(DUT_Data_Sheet[col + str(4)].value)  # LLimit
            test_info.upper_limit = float(DUT_Data_Sheet[col + str(3)].value)  # HLimit
            test_info.unit = DUT_Data_Sheet[col + str(5)].value  # Unit
            values = []
            x = []
            y = []
            temp_passed = []
            total_wafer_run_passed = []
            passed_run = []
            for row_number in range(6, max_row + 1):
                coord = DUT_Data_Sheet["J" + str(row_number)].value
                x.append(int(coord.split()[0][1:-1]))
                y.append(int(coord.split()[1][0:-1]))
                is_failed = DUT_Data_Sheet["K" + str(row_number)].value
                if is_failed == "Failed - 0x08":
                    total_wafer_run_passed.append(False)
                    temp_passed.append(False)
                    passed_run.append(False)
                    # if filter_type == FilterType.PASSED_RUN:
                    #     continue
                else:
                    total_wafer_run_passed.append(True)
                    temp_passed.append(True)
                    passed_run.append(True)
                cell_value = DUT_Data_Sheet[col + str(row_number)].value
                try:
                    values.append(float(cell_value))
                    if test_info.lower_limit <= cell_value <= test_info.upper_limit:
                        temp_passed[-1] = True
                except:
                    values.append(np.nan)
                    temp_passed[-1] = None

            test_info.x = np.array(x)
            test_info.y = np.array(y)
            test_info.is_passed = np.array(temp_passed, dtype=object)
            total_passed_array = np.array(total_wafer_run_passed)
            passed_array = np.array(passed_run, dtype=bool)


            # reverse y and shift x and y to coordinate 0, 0
            x_min = min(test_info.x)
            y_shift = max(test_info.y)
            test_info.x = test_info.x + abs(x_min)
            # test_info.y = y_shift - test_info.y
            # Thread(
            #     target=generate_wafer_map,
            #     args=(test_info.x, test_info.y, y_shift, test_info.is_passed, str(test_id)+"_wafermap.png",
            #           from_enum_to_string(filter_type)+"_wafer_map/"),
            #     daemon=True
            # ).start()

            if not total_wafer_run:
                generate_wafer_map(test_info.x, test_info.y, y_shift, total_passed_array,
                                   f'{wafer_run_info.family_id} Wafer Run',
                                   f'{wafer_run_info.family_id}_Wafer_Run.png',
                                   f'{wafer_run_info.family_id} wafer map/')

            total_wafer_run = True

            title_ = f'Test ID: {test_id} Wafer Map' if (filter_type == FilterType.ALL_RUN or
                                                         filter_type == FilterType.PASSED_RUN) \
                else f'Test ID: {test_id} before {title} filter'
            generate_wafer_map(test_info.x, test_info.y, y_shift, test_info.is_passed, title_,
                               str(test_id)+"_before_wafermap.png",
                               title + "_wafer_map/")

            test_info.values = np.array(values)
            value_array_backup = copy.copy(test_info.values)
            # print(test_info.values, len(test_info.values))
            if filter_type == FilterType.PASSED_RUN:
                # passed_index = np.where(passed_array == 1)
                # print(passed_index, len(passed_index))
                test_info.values = test_info.values[passed_array]
            # print(test_info.values)

            no_nan_mask = ~np.isnan(value_array_backup)
            # nan_mask = np.isnan(value_array_backup)
            test_info.total_run = len(value_array_backup[no_nan_mask])
            if filter_type == FilterType.PASSED_RUN:
                test_info.dropped = (int(wafer_run_info.all_runs) - int(wafer_run_info.passed_runs) -
                                     len(value_array_backup[~no_nan_mask]))
                # test_info.total_run = int(wafer_run_info.all_runs)
                # test_info.dropped = int(wafer_run_info.all_runs) - int(wafer_run_info.passed_runs)
                # print(test_info.dropped, test_info.total_run)
            elif filter_type == FilterType.ALL_RUN:
                # test_info.total_run = int(wafer_run_info.all_runs)
                test_info.dropped = 0
            test_info.generate_statistics()

            # select data based on filter
            # for filter type three sigma
            # if test_info.unit != "HEX" and test_info.stats.std > 0.001:

            if test_info.stats.std > 0.0001:
                mask = np.array([], dtype=bool)
                if filter_type != FilterType.ALL_RUN and filter_type != FilterType.PASSED_RUN:
                    # filter one sigma
                    if filter_type == FilterType.ONE_SIGMA:
                        mask = (((test_info.stats.mean - test_info.stats.std) <= test_info.values) &
                                (test_info.values <= (test_info.stats.mean + test_info.stats.std)))

                    # filter two sigma
                    elif filter_type == FilterType.TWO_SIGMA:
                        mask = (((test_info.stats.mean - 2 * test_info.stats.std) <= test_info.values) &
                                (test_info.values <= (test_info.stats.mean + 2 * test_info.stats.std)))

                    # filter three sigma
                    elif filter_type == FilterType.THREE_SIGMA:
                        mask = (((test_info.stats.mean - 3 * test_info.stats.std) <= test_info.values) &
                                (test_info.values <= (test_info.stats.mean + 3 * test_info.stats.std)))

                    # filter six sigma
                    elif filter_type == FilterType.SIX_SIGMA:
                        mask = (((test_info.stats.mean - 6 * test_info.stats.std) <= test_info.values) &
                                (test_info.values <= (test_info.stats.mean + 6 * test_info.stats.std)))

                    # update pass array and value array
                    index_array, = np.where(~mask)
                    for i in index_array:
                        if test_info.is_passed[i]:
                            test_info.is_passed[i] = np.nan
                    test_info.values = test_info.values[mask]
                    test_info.dropped = test_info.total_run - len(test_info.values)

                    # recalculate stats result after filter
                    if len(test_info.values) > 1:
                        test_info.generate_statistics()

                    # generate wafer map after filter
                    generate_wafer_map(test_info.x, test_info.y, y_shift, test_info.is_passed, f'Test ID: {test_id} '
                                                                                               f' after {title} filter',
                                       str(test_id) + "_after_wafermap.png",
                                       title + "_wafer_map/")

                    # combine two wafer map into one
                    images = [Image.open(title + "_wafer_map/"+str(test_id) + "_before_wafermap.png"),
                              Image.open(title + "_wafer_map/"+str(test_id) + "_after_wafermap.png")]
                    widths, heights = zip(*(i.size for i in images))

                    total_width = sum(widths)
                    max_height = max(heights)

                    new_im = Image.new('RGB', (total_width, max_height))
                    x_offset = 0
                    for im in images:
                        new_im.paste(im, (x_offset, 0))
                        x_offset += im.size[0]
                    # save to local dir
                    new_im.save(title + "_wafer_map/"+str(test_id) + "_before_after_wafermap.png")

            if test_id not in test_dict:
                test_dict[test_id] = test_info

    cpk_9999 = []

    for key, value in test_dict.items():
        # print(f'{value.test_description}: cp: {value.stats.cp}, cpk: {value.stats.cpk}')
        if value.stats.cp == -9999.999:
            cpk_9999.append([key, test_dict[key]])

    sorted_ = {k: v for k, v in sorted(test_dict.items(), key=lambda item: item[1].stats.cpk)}
    cpk_first_n = list(sorted_.items())[:40]

    statistics_report = "Statistics Report"

    pdf = FPDF()

    # add title
    pdf.add_page()
    pdf.set_font("helvetica", size=40, style="IB")
    pdf.set_text_color(255, 0, 0)
    pdf.cell(195, 70, txt=statistics_report, ln=1, align="C")

    pdf.set_font("helvetica", size=20, style="IB")
    pdf.set_text_color(0, 0, 250)
    pdf.cell(195, 10, txt=wafer_run_info.family_id, ln=1, align="C")
    #
    pdf.set_font("helvetica", size=16, style="B")
    pdf.set_text_color(0, 0, 0)
    pdf.cell(10, 10, txt=str(" " * 40 + str(wafer_run_info.product_id)), ln=1, align="L")
    #
    pdf.set_font("helvetica", size=16, style="B")
    pdf.set_text_color(0, 0, 0)
    pdf.cell(10, 10, txt=" " * 40 + wafer_run_info.start_time, ln=1, align="L")
    #
    pdf.set_font("helvetica", size=16, style="B")
    pdf.set_text_color(0, 0, 0)
    pdf.cell(10, 10, txt=" " * 108 + "OS: " + wafer_run_info.os, ln=1, align="C")
    #
    pdf.set_font("helvetica", size=16, style="B")
    pdf.set_text_color(0, 0, 0)
    pdf.cell(10, 10, txt=" " * 137 + "Tester: " + wafer_run_info.tester, ln=1, align="C")
    #
    pdf.set_font("helvetica", size=16, style="B")
    pdf.set_text_color(0, 0, 0)
    pdf.cell(10, 10, txt=" " * 96 + "Lot: " + str(wafer_run_info.lot), ln=1, align="C")
    #
    pdf.set_font("helvetica", size=16, style="B")
    pdf.set_text_color(0, 0, 0)
    pdf.cell(10, 10, txt=" " * 88 + "Sites: " + str(wafer_run_info.sites), ln=1, align="C")
    #
    pdf.set_font("helvetica", size=16, style="B")
    pdf.set_text_color(0, 0, 0)

    num_runs = wafer_run_info.passed_runs if filter_type == FilterType.PASSED_RUN else wafer_run_info.all_runs
    run_info = "All Runs: " if filter_type != FilterType.PASSED_RUN else "Passed Runs: "
    spacing = 100 if filter_type != FilterType.PASSED_RUN else 108
    pdf.cell(10, 10, txt=" " * spacing + run_info + str(num_runs), ln=1, align="C")

    pdf.image(f'{wafer_run_info.family_id} wafer map/{wafer_run_info.family_id}_Wafer_Run.png',
              x=45, y=170, w=120, h=110)

    table_data = [["TNUM", "SITE", "RANGE", "STD", "CP", "CPK", "UNITS", "TNAME"]]
    link_dic = {}
    for k, v in cpk_first_n:
        temp = []
        temp.append(str(k))
        temp.append(1)
        temp.append(round(v.stats.range, 3))
        temp.append(round(v.stats.std, 3))
        temp.append(round(v.stats.cp, 3))
        temp.append(round(v.stats.cpk, 3))
        temp.append(v.unit)
        temp.append(v.test_description)
        link_dic[v.test_description] = pdf.add_link()
        table_data.append(temp)

    pdf.add_page()
    pdf.set_font("helvetica", size=20, style="B")
    pdf.set_text_color(0, 0, 0)
    pdf.cell(10, 10, txt=" " * 85 + "Cpk Summary", ln=1, align="C")
    #
    pdf.ln(5)
    pdf.set_font('Helvetica', 'b', 8)
    pdf.set_text_color(0, 0, 0)
    row = 0
    col = 0

    for row_index, data_row in enumerate(table_data):
        for column_index, data in enumerate(data_row):
            if column_index != 7:
                if row_index == 0 or column_index != 0:
                    pdf.cell(w=18, h=6, txt=str(data), border=1, ln=0, align="R")
                else:
                    pdf.cell(w=18, h=6, txt=str(data), border=1, ln=0, align="R", link=link_dic[data_row[7]])
            else:
                if row_index == 0:
                    pdf.cell(w=54, h=6, txt=str(data), border=1, ln=0, align="R")
                else:
                    pdf.cell(w=54, h=6, txt=str(data), border=1, ln=0, align="R", link=link_dic[data])
        pdf.ln()

    # add trend plot and histogram
    plt.style.use('ggplot')
    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(10, 8), constrained_layout=True)
    # fig.subplots_adjust(top=0.9)
    # fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(8, 6))

    trend_title = "Trend Plot"
    histogram = "Histogram"
    index = 0
    load = len(test_dict)

    for test_id, test_info in test_dict.items():
        # print(test_id)
        test_description = test_info.test_description
        min_ = round(test_info.stats.min, 3)
        max_ = round(test_info.stats.max, 3)
        mean = round(test_info.stats.mean, 3)
        median = round(test_info.stats.median, 3)
        range_ = round(test_info.stats.range, 3)
        std = round(test_info.stats.std, 3)
        cp = round(test_info.stats.cp, 3)
        cpk = round(test_info.stats.cpk, 3)
        unit = test_info.unit
        l_lim = test_info.lower_limit
        h_lim = test_info.upper_limit
        number_points = len(test_info.values)
        number_array = np.arange(1, number_points + 1)
        value_array = test_info.values
        ax1.scatter(number_array, value_array, c="blue", s=2, marker="o")
        ax1.plot(number_array, value_array, "r-", linewidth=1)
        ax1.plot(number_array, np.repeat(h_lim, number_points), linewidth=2, linestyle="--", label=f"HLIM: {h_lim}")
        ax1.plot(number_array, np.repeat(l_lim, number_points), linewidth=2, linestyle="--", label=f"LLIM: {l_lim}")
        ax1.legend(loc="upper right")
        ax1.set_xlabel("runs")
        limit_gap = ((h_lim - l_lim) / 5.) * 0.2
        ax1.set_ylabel(unit)
        ax1.set_ylim(l_lim - limit_gap, h_lim + limit_gap)
        ax1.set_title(trend_title)

        ax2.axvline(x=l_lim, color="blue", linestyle="-", label=f"ll={l_lim}")
        ax2.axvline(x=h_lim, color="blue", linestyle="-", label=f"ul={h_lim}")
        ax2.axvline(x=median, color="grey", linestyle="--", label=f"median={median}")

        if filter_type == FilterType.ONE_SIGMA or filter_type == FilterType.PASSED_RUN or filter_type == FilterType.ALL_RUN:
            ax2.axvline(x=mean - std, color="green", linestyle="-",
                        label=f"{from_enum_to_string(FilterType.ONE_SIGMA)} ll={round(mean - std, 3)}")
            ax2.axvline(x=mean + std, color="green", linestyle="-",
                        label=f"{from_enum_to_string(FilterType.ONE_SIGMA)} ul={round(mean + std, 3)}")
        elif filter_type == FilterType.TWO_SIGMA:
            ax2.axvline(x=mean - 2 * std, color="green", linestyle="-",
                        label=f"{title} ll={round(mean- 2 * std, 3)}")
            ax2.axvline(x=mean + 2 * std, color="green", linestyle="-",
                        label=f"{title} ul={round(mean + 2 * std, 3)}")
        elif filter_type == FilterType.THREE_SIGMA:
            ax2.axvline(x=mean - 3 * std, color="green", linestyle="-",
                        label=f"{title} ll={round(mean - 3 * std, 3)}")
            ax2.axvline(x=mean + 3 * std, color="green", linestyle="-",
                        label=f"{title} ul={round(mean + 3 * std, 3)}")
        elif filter_type == FilterType.SIX_SIGMA:
            ax2.axvline(x=mean - 6 * std, color="green", linestyle="-",
                        label=f"{title} ll={round(mean - 6 * std, 3)}")
            ax2.axvline(x=mean + 6 * std, color="green", linestyle="-",
                        label=f"{title} ul={round(mean + 6 * std, 3)}")

        ax2.hist(value_array, bins=10, color="skyblue", edgecolor="black", align='mid')

        ax2.set_xlabel(unit)
        ax2.tick_params(axis='x', labelrotation=45)
        ax2.xaxis.set_major_formatter(FormatStrFormatter('%.2f'))
        ax2.set_title(histogram)
        ax2.legend(loc="upper right")
        plt.tight_layout()
        # fig.suptitle(str(test_id) + ": " + test_info.test_description, fontsize="x-large")
        fig.suptitle(f"Test {test_id} - {test_info.test_description}\n"
                     f"fallout = {test_info.dropped} out of {test_info.total_run} "
                     f"({round((test_info.dropped/test_info.total_run)*100, 2)}%)\n\u03BC = {mean} \u03C3 = {std} cpk = {round(cpk, 1)}\n"
                     f"dropped = {test_info.dropped}",
                     fontsize=14)
        temp = ""
        if filter_type == FilterType.PASSED_RUN:
            temp = "passed_runs_"
        elif filter_type == FilterType.ALL_RUN:
            temp = "all_runs_"
        elif filter_type == FilterType.ONE_SIGMA:
            temp = "1_sigma_runs_"
        elif filter_type == FilterType.TWO_SIGMA:
            temp = "2_sigma_runs_"
        elif filter_type == FilterType.THREE_SIGMA:
            temp = "3_sigma_runs_"
        else:
            temp = "6_sigma_runs_"

        directory = temp + "plots/"
        if not os.path.exists(directory):
            os.makedirs(directory, exist_ok=True)

        wafer_map_file_name = f"{title}_wafer_map/{test_id}_before_after_wafermap.png"
        plot_file_name = directory + str(test_id) + ".png"
        fig.savefig(plot_file_name)
        ax1.cla()
        ax2.cla()

        pdf.add_page()
        if test_description in link_dic:
            pdf.set_link(link_dic[test_description])
        pdf.add_font(fname='DejaVuSansCondensed.ttf')
        pdf.set_font("helvetica", style="B", size=15)
        pdf.set_text_color(0, 0, 0)
        pdf.cell(180, 10, txt=f"Test: {test_id}", ln=1, align="C")
        pdf.set_font("helvetica", style="B", size=20)
        pdf.set_text_color(0, 0, 0)
        pdf.cell(180, 10, txt=f"{test_description}", ln=1, align="C")
        pdf.ln(3)
        pdf.set_font("DejaVuSansCondensed", size=10)
        pdf.set_text_color(0, 0, 0)
        parameter_table = [["SITE", "FILTER", "MIN", "MEAN", "MAX", "RANGE", "STD", "CP", "CPK", "UNITS"],
                           [wafer_run_info.sites, title, min_, mean, max_, range_, std, cp, cpk, unit]]

        for data_row in parameter_table:
            for row_index, data in enumerate(data_row):
                pdf.cell(w=19, h=6, txt=" " * 4 + str(data), border=1, ln=0, align="R")
            pdf.ln()

        # insert wafer map
        if filter_type != FilterType.ALL_RUN and filter_type != FilterType.PASSED_RUN:
            pdf.image(f"{wafer_map_file_name}", x=6, y=55, w=200, h=95)
        else:
            pdf.image(f"{title}_wafer_map/{test_id}_before_wafermap.png", x=45, y=55, w=120, h=110)

        # insert plot
        pdf.image(f"{plot_file_name}", x=0, y=170, w=200, h=100)
        index += 1
        progress_val = index/float(load)*100
        progress_bar["value"] = progress_val
        if progress_label is not None:
            progress_label.config(text=f"{progress_val:.0f}%")
        progress_bar.update_idletasks()

    pdf.output(export_file)

    if "src_path" not in src_dict:
        src_dict["src_path"] = []
    src_dict["src_path"].append(export_file)

    # add outline of pdf
    pdf_ = Pdf.open(export_file)
    outline_list = []
    with pdf_.open_outline() as outline:
        oi = OutlineItem("Cover page", 0)
        outline_list.append(oi)
        oi = OutlineItem("Cpk summary", 1)
        outline_list.append(oi)
        page_index = 1
        for test_id, test_info in test_dict.items():
            page_index += 1
            oi = OutlineItem(str(test_id) + ": " + test_info.test_description, page_index)
            outline_list.append(oi)
        outline.root.extend(outline_list)

    # add link page
    if not os.path.exists("../wafer run stats report tool/statistics report/"):
        os.makedirs("../wafer run stats report tool/statistics report/", exist_ok=True)
    pdf_.save(os.getcwd() + "/statistics report/" + os.path.basename(export_file))

    messagebox.showinfo("Completion!", "Generate stats report successfully!")