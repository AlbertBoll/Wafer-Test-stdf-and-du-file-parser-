import os.path
import sys

import numpy as np

lib_dir = os.path.join(os.getcwd(), "ChartDirector\lib")
sys.path.insert(0, lib_dir)
from pychartdir import *
from openpyxl import load_workbook

DUT_SUMMARY = "DUT Summary"


def generate_stats_report_(import_file, filter_type, export_file, progress_bar, progress_label=None, src_dict={}):
    wook_book = load_workbook(import_file)
    DUT_Data_Sheet = wook_book[DUT_SUMMARY]
    max_row = DUT_Data_Sheet.max_row
    max_column = DUT_Data_Sheet.max_column


class die_run:
    def __init__(self):
        self.coord = coordinate()
        self.is_passed = False

class coordinate:
    def __init__(self, x = 0, y = 0):
        self.x = x
        self.y = y


wafer_dies_info = []


def generate_wafer_map(x_coord, y_coord, shift, is_pass, title, output_file_name, output_file_dir="wafer map/"):
    diameter = shift + 1
    # The random data array are for a square grid of 20 x 20 cells
    r = RanSeries(2)
    zData = r.get2DSeries(diameter, diameter, 0, 2)

    for i in range(0, len(zData)):
        i_ = i % diameter
        j_ = i // diameter
        is_found = False
        for index, (_x, _y) in enumerate(zip(x_coord, y_coord)):
            if _x == i_ and _y == j_:
                if is_pass[index] is not None and not np.isnan(is_pass[index]):
                    val = 1 if is_pass[index] else 0
                    zData[i] = val
                elif is_pass[index] is None:
                    zData[i] = 0.5
                elif np.isnan(is_pass[index]):
                    zData[i] = 0.8
                is_found = True
                break
        if not is_found:
            zData[i] = NoValue

    c = XYChart(520, 480)
    # Add a title the chart with 15pt Arial Bold font
    c.addTitle(title, "Arial Bold", 15)

    # Set the plotarea at (50, 40) and of size 400 x 400 pixels. Set the backgound and border to
    # transparent. Set both horizontal and vertical grid lines to light grey. (0xdddddd)
    p = c.setPlotArea(50, 40, 400, 400, -1, -1, Transparent, 0xdddddd, 0xdddddd)

    # Create a discrete heat map with diameter x diameter cells
    layer = c.addDiscreteHeatMapLayer(zData, diameter)

    # Set the x-axis scale. Use 8pt Arial Bold font. Set axis color to transparent, so only the labels
    # visible. Set 0.5 offset to position the labels in between the grid lines.

    c.xAxis().setLinearScale(0, diameter, 1)
    c.xAxis().setLabelStyle("Arial Bold", 8)
    c.xAxis().setColors(Transparent, TextColor)
    c.xAxis().setLabelOffset(0.5)
    c.setXAxisOnTop()

    # Set the y-axis scale. Use 8pt Arial Bold font. Set axis color to transparent, so only the labels
    # visible. Set 0.5 offset to position the labels in between the grid lines.
    c.yAxis().setLinearScale(0, diameter, 1)
    c.yAxis().setLabelStyle("Arial Bold", 8)
    c.yAxis().setColors(Transparent, TextColor)
    c.yAxis().setLabelOffset(0.5)
    c.yAxis().setReverse()

    # Position the color axis 20 pixels to the right of the plot area and of the same height as the plot
    # area. Put the labels on the right side of the color axis. Use 8pt Arial Bold font for the labels.
    # cAxis = layer.setColorAxis(p.getRightX() + 20, p.getTopY(), TopLeft, p.getHeight(), Right)
    # cAxis.setLabelStyle("Arial Bold", 8)

    colorLabels = ["Failed", "Aborted", "Filtered", "Passed"]
    colorScale = [0, 0xff0000, 0.5, 0xaaaaaa, 0.8, 0x9600ff, 1, 0x00ff00, 2]
    # colorScale = [0, 0xff0000, 1, 0xff8800, 3, 0x4488cc, 7, 0x99ccff, 9, 0x00ff00, 10]

    layer.colorAxis().setColorScale(colorScale)

    b = c.addLegend(p.getRightX() + 7, p.getTopY(), 1, "Arial Bold", 7)
    b.setBackground(Transparent, Transparent)
    b.setKeySize(15, 15)
    b.setKeySpacing(0, 8)

    # Add the color scale label to the legend box
    for i in range(len(colorLabels) - 1, -1, -1):
        b.addKey(colorLabels[i], int(colorScale[i * 2 + 1]))

    # Output the chart
    # if not os.path.exists(output_file_dir):
    #     os.makedirs(output_file_dir, exist_ok=True)
    c.makeChart(output_file_dir + output_file_name)